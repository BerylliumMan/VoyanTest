from __future__ import annotations

import io
import logging
from typing import Any, Dict, List

from fastapi import APIRouter, File, HTTPException, UploadFile, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app import crud, models, db_models
from app.auth import require_admin, require_project_access, get_user_project_filter, get_current_user
from app.database import get_async_db

logger = logging.getLogger(__name__)

router = APIRouter()


@router.post("/", response_model=models.TestCase)
async def create_test_case(case: models.TestCaseCreate, user=Depends(get_current_user), db: AsyncSession = Depends(get_async_db)) -> models.TestCase:
    allowed_ids = get_user_project_filter(user)
    if allowed_ids is not None and case.project_id not in allowed_ids:
        raise HTTPException(status_code=403, detail="无权访问该项目")
    """
    创建带有步骤的新测试用例。
    """
    db_project = await crud.get_project(db, case.project_id)
    if db_project is None:
        raise HTTPException(status_code=404, detail=f"Project with id {case.project_id} not found")

    try:
        return await crud.create_test_case(db, case)
    except Exception as e:
        logger.exception("创建测试用例失败")
        raise HTTPException(status_code=400, detail="创建测试用例失败，请检查输入数据")


@router.get("/search", response_model=models.TestCasePage)
async def search_test_cases(project_id: int, q: str = "", page: int = 1, size: int = 20, user=Depends(get_current_user), db: AsyncSession = Depends(get_async_db)) -> dict:
    """
    搜索测试用例（按名称和描述）。
    """
    allowed_ids = get_user_project_filter(user)
    if allowed_ids is not None and project_id not in allowed_ids:
        raise HTTPException(status_code=404, detail="Project not found")
    result = await crud.search_test_cases(db, project_id, q, page, size)
    return {
        "items": result["items"],
        "total_items": result["total_items"],
        "page": page,
        "size": size,
    }


@router.get("/init-cases", response_model=List[models.TestCase])
async def list_init_test_cases(project_id: int, user=Depends(get_current_user), db: AsyncSession = Depends(get_async_db)) -> list[models.TestCase]:
    """获取项目下所有标记为初始化的测试用例"""
    allowed_ids = get_user_project_filter(user)
    if allowed_ids is not None and project_id not in allowed_ids:
        raise HTTPException(status_code=404, detail="Project not found")
    return await crud.get_init_test_cases(db, project_id)


@router.get("/export")
async def export_test_cases(
    project_id: int,
    user=Depends(get_current_user),
    db: AsyncSession = Depends(get_async_db),
):
    """导出项目测试用例为 xlsx（格式对齐 AI 生成下载格式）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    cases = await crud.get_all_test_cases_for_project(db, project_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 表头样式（对齐 gen/history.py 格式）
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    cell_align = Alignment(vertical="top", wrap_text=True)

    headers = ["用例ID", "所属模块", "标题", "前置条件", "测试步骤", "预期结果", "优先级"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, tc in enumerate(cases, 2):
        # 步骤合并为文本块（每行 "序号. 描述 → 预期结果"）
        steps_str = ""
        expected_str = ""
        if tc.steps:
            lines = []
            exp_lines = []
            for s in tc.steps:
                lines.append(f"{s.step_order}. {s.description}")
                exp_lines.append(f"{s.step_order}. {s.parsed_result or ''}")
            steps_str = "\n".join(lines)
            expected_str = "\n".join(exp_lines)

        module_name = tc.module.name if tc.module else ""

        values = [
            tc.project_case_number,
            module_name,
            tc.name,
            tc.description or "",
            steps_str,
            expected_str,
            tc.priority or "medium",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = cell_align
            cell.border = thin_border

    # 列宽（对齐 gen/history.py 格式）
    widths = [10, 16, 30, 24, 40, 40, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=testcases_project_{project_id}.xlsx"},
    )


@router.get("/{case_id}", response_model=models.TestCase)
async def get_test_case(case_id: int, user=Depends(get_current_user), db: AsyncSession = Depends(get_async_db)) -> models.TestCase:
    """
    通过其ID检索单个测试用例，包括其步骤。
    """
    db_case = await crud.get_test_case(db, case_id)
    if db_case is None:
        raise HTTPException(status_code=404, detail="Test case not found")
    # 验证项目访问权限
    allowed_ids = get_user_project_filter(user)
    if allowed_ids is not None and db_case.project_id not in allowed_ids:
        raise HTTPException(status_code=404, detail="Test case not found")
    return db_case


@router.delete("/{case_id}")
async def delete_test_case(case_id: int, user=Depends(get_current_user), db: AsyncSession = Depends(get_async_db)) -> dict[str, str] | None:
    """
    删除测试用例及其步骤。
    """
    db_case = await crud.get_test_case(db, case_id)
    if db_case is None:
        raise HTTPException(status_code=404, detail="Test case not found")
    allowed_ids = get_user_project_filter(user)
    if allowed_ids is not None and db_case.project_id not in allowed_ids:
        raise HTTPException(status_code=403, detail="无权访问该项目")

    try:
        return await crud.delete_test_case(db, case_id)
    except Exception as e:
        logger.exception("删除测试用例失败 case_id=%s", case_id)
        raise HTTPException(status_code=500, detail="删除测试用例时发生内部错误")


@router.put("/{case_id}", response_model=models.TestCase)
async def update_test_case(case_id: int, case: models.TestCaseUpdate, user=Depends(get_current_user), db: AsyncSession = Depends(get_async_db)) -> models.TestCase:
    """
    更新测试用例，包括其步骤。
    """
    db_case = await crud.get_test_case(db, case_id)
    if db_case is None:
        raise HTTPException(status_code=404, detail="Test case not found")
    allowed_ids = get_user_project_filter(user)
    if allowed_ids is not None and db_case.project_id not in allowed_ids:
        raise HTTPException(status_code=403, detail="无权访问该项目")

    try:
        return await crud.update_test_case(db, case_id, case)
    except Exception as e:
        logger.exception("更新测试用例失败 case_id=%s", case_id)
        raise HTTPException(status_code=500, detail="更新测试用例时发生内部错误")


@router.get("/module/{module_id}/testcases", response_model=models.TestCasePage)
async def get_module_test_cases(module_id: int, page: int = 1, size: int = 20, user=Depends(get_current_user), db: AsyncSession = Depends(get_async_db)) -> dict:
    """
    检索特定模块的所有测试用例，并分页。
    """
    db_module = await crud.get_module(db, module_id)
    if db_module is None:
        raise HTTPException(status_code=404, detail="Module not found")
    # 验证项目访问权限
    allowed_ids = get_user_project_filter(user)
    if allowed_ids is not None and db_module.project_id not in allowed_ids:
        raise HTTPException(status_code=404, detail="Module not found")

    paginated_data = await crud.get_all_test_cases_for_module_paginated(db, module_id, page, size)
    return {
        "items": paginated_data["items"],
        "total_items": paginated_data["total_items"],
        "page": page,
        "size": size,
    }


@router.put("/{case_id}/toggle-init", response_model=models.TestCase)
async def toggle_test_case_init(case_id: int, body: Dict[str, Any], user=Depends(get_current_user), db: AsyncSession = Depends(get_async_db)) -> models.TestCase:
    """切换测试用例的初始化标记"""
    db_case = await crud.get_test_case(db, case_id)
    if db_case is None:
        raise HTTPException(status_code=404, detail="Test case not found")
    allowed_ids = get_user_project_filter(user)
    if allowed_ids is not None and db_case.project_id not in allowed_ids:
        raise HTTPException(status_code=403, detail="无权访问该项目")
    is_init = body.get("is_init", False)
    db_case = await crud.update_test_case_is_init(db, case_id, is_init)
    if db_case is None:
        raise HTTPException(status_code=404, detail="Test case not found")
    return db_case


@router.get("/project/{project_id}/testcases", response_model=models.TestCasePage)
async def get_project_test_cases(project_id: int, page: int = 1, size: int = 20, user=Depends(get_current_user), db: AsyncSession = Depends(get_async_db)) -> dict:
    """
    检索特定项目的所有测试用例，并分页。
    """
    allowed_ids = get_user_project_filter(user)
    if allowed_ids is not None and project_id not in allowed_ids:
        raise HTTPException(status_code=404, detail="Project not found")
    db_project = await crud.get_project(db, project_id)
    if db_project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    paginated_data = await crud.get_all_test_cases_for_project_paginated(db, project_id, page, size)
    return {
        "items": paginated_data["items"],
        "total_items": paginated_data["total_items"],
        "page": page,
        "size": size,
    }


@router.post("/import")
async def import_test_cases(
    project_id: int,
    file: UploadFile = File(...),
    user=Depends(get_current_user),
    db: AsyncSession = Depends(get_async_db),
) -> dict:
    """从 xlsx 文件导入测试用例。"""
    allowed_ids = get_user_project_filter(user)
    if allowed_ids is not None and project_id not in allowed_ids:
        raise HTTPException(status_code=403, detail="无权访问该项目")
    from openpyxl import load_workbook

    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件")

    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents))
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=400, detail="工作表为空")

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    created = 0
    errors = []

    # 从表头判断格式（新格式含"测试步骤"/"前置条件"，旧格式含"步骤序号"/"步骤描述"）
    header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    has_header = header_cells and any(str(c or "").strip() in ("前置条件", "测试步骤") for c in header_cells[0])

    for row in rows:
        if not any(str(c or "").strip() for c in row):
            continue

        if has_header:
            # 新格式：用例ID, 所属模块, 标题, 前置条件, 测试步骤, 预期结果, 优先级
            cells = [str(c or "").strip() for c in row]
            case_id, module_name, title, preconditions, steps_text, expected_text, priority = (
                (cells + [""] * 7)[:7]
            )
            if not title:
                continue

            try:
                # 解析合并的步骤文本
                steps_payload = []
                if steps_text:
                    for line in steps_text.split("\n"):
                        line = line.strip()
                        if not line:
                            continue
                        desc = line
                        exp = ""
                        if " → " in line:
                            desc, exp = line.split(" → ", 1)
                        # 去掉序号前缀 "1. "
                        if ". " in desc and desc[0].isdigit():
                            desc = desc.split(". ", 1)[1]
                        steps_payload.append(models.TestStepCreatePayload(
                            step_order=len(steps_payload) + 1,
                            description=desc.strip(),
                            parsed_result=exp.strip() or None,
                        ))

                await crud.create_test_case(db, models.TestCaseCreate(
                    project_id=project_id, name=title,
                    description=preconditions or "",
                    steps=steps_payload, priority=priority or "medium",
                ))
                created += 1
            except Exception as e:
                errors.append(f"创建用例「{title}」失败: {e}")
        else:
            # 旧格式（向后兼容）：用例名称, 模块, 步骤序号, 步骤描述, 预期结果, 优先级, 标签
            cells = [str(c or "").strip() for c in row]
            name, _module, _so, desc, expected, priority, tags = (cells + [""] * 7)[:7]
            if not name:
                continue
            try:
                steps_payload = []
                if desc:
                    steps_payload.append(models.TestStepCreatePayload(
                        step_order=1, description=desc, parsed_result=expected or None,
                    ))
                await crud.create_test_case(db, models.TestCaseCreate(
                    project_id=project_id, name=name,
                    steps=steps_payload, priority=priority or "medium", tags=tags or "",
                ))
                created += 1
            except Exception as e:
                errors.append(f"创建用例「{name}」失败: {e}")

    return {"created": created, "errors": errors}
