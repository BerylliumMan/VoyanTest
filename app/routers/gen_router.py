"""API router for AI test case generation."""
import os
import logging
import threading
import json
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query, Response
from sqlalchemy.orm import Session
from typing import Optional, List
from pydantic import BaseModel

from ..database import get_db
from .. import db_models
from ..auth import get_current_user

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/gen", tags=["用例生成"])

# In-memory session store (survives within process lifetime)
_sessions: dict = {}
_lock = threading.Lock()


class GenStatusResponse(BaseModel):
    session_id: str
    status: str  # pending / analyzing / completed / failed
    filename: str = ""
    error_message: str = ""
    functional_points_count: int = 0
    test_cases_count: int = 0


class GenPreviewItem(BaseModel):
    test_case_id: str
    module: str
    title: str
    preconditions: str
    test_steps: str
    expected_result: str
    priority: str
    selected: bool = True


class GenPreviewResponse(BaseModel):
    session_id: str
    functional_points: list[dict]
    test_cases: list[GenPreviewItem]


class GenImportRequest(BaseModel):
    session_id: str
    project_id: int
    selected_ids: list[str] | None = None  # None = import all


class GenImportResponse(BaseModel):
    imported_count: int
    test_case_ids: list[int]


class GenHistoryItem(BaseModel):
    id: str
    filename: str
    filenames: list[str]
    project_id: Optional[int] = None
    project_name: str = ""
    project_description: str
    status: str
    error_message: str
    functional_points_count: int
    test_cases_count: int
    imported_count: int
    created_at: datetime
    completed_at: Optional[datetime]


class GenHistoryListResponse(BaseModel):
    items: list[GenHistoryItem]
    total: int


ALLOWED_EXTENSIONS = {".docx", ".md", ".png", ".jpg", ".jpeg", ".pdf"}
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB

# 魔术字节签名对照表
_MAGIC_SIGNATURES: dict[str, list[bytes]] = {
    ".png": [b"\x89PNG\r\n\x1a\n"],
    ".jpg": [b"\xff\xd8\xff"],
    ".jpeg": [b"\xff\xd8\xff"],
    ".pdf": [b"%PDF-"],
    ".docx": [b"PK\x03\x04"],
}

# .md 无固定魔术字节，跳过二进制检查


def _check_extension(filename: str):
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, f"不支持的文件类型：'{filename}' (ext='{ext}')，仅支持 .docx/.md/.png/.jpg/.jpeg/.pdf")
    return ext


def _check_magic_bytes(data: bytes, ext: str):
    """校验文件头魔术字节，阻止伪装扩展名的攻击文件。"""
    sigs = _MAGIC_SIGNATURES.get(ext, [])
    if not sigs:
        return  # .md 等无固定签名的类型跳过
    if not any(data.startswith(sig) for sig in sigs):
        raise HTTPException(400, f"文件内容与扩展名 '{ext}' 不匹配，疑似伪装文件")


@router.post("/upload")
async def upload_and_analyze(
    files: List[UploadFile] = File(...),
    project_description: str = Form(""),
    project_id: Optional[int] = Form(None),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Upload document(s) and start AI analysis to generate test cases."""
    if not files:
        raise HTTPException(400, "请上传至少一个文件")

    filenames = [f.filename or f"file_{i}" for i, f in enumerate(files)]

    from io import BytesIO
    file_contents = []
    for f in files:
        ext = _check_extension(f.filename or "")
        content = await f.read()
        _check_magic_bytes(content, ext)
        file_contents.append(BytesIO(content))

    from app.gen.models import AnalysisSession
    import uuid
    session_id = str(uuid.uuid4())
    session = AnalysisSession(
        session_id=session_id,
        filename=filenames[0] if filenames else "unknown",
        filenames=filenames,
        project_description=project_description,
        status="analyzing",
    )

    with _lock:
        _sessions[session_id] = session

    db_record = db_models.GenSession(
        id=session_id,
        filename=filenames[0] if filenames else "unknown",
        filenames=json.dumps(filenames),
        project_id=project_id,
        project_description=project_description,
        status="analyzing",
    )
    db.add(db_record)
    db.commit()

    def _run_full_analysis():
        try:
            from app.gen.analyzer import extract_multi_file_content, two_phase_analyze, get_default_prompts
            from app.database import SessionLocal

            combined_text, _, warnings = extract_multi_file_content(
                file_contents, filenames
            )

            thread_db = SessionLocal()
            try:
                defaults = get_default_prompts()
                prompt_rows = thread_db.query(db_models.PromptTemplate).all()
                prompts = {}
                for row in prompt_rows:
                    if row.is_custom and row.template_key in defaults:
                        prompts[row.template_key] = {"content": row.template_content}
            finally:
                thread_db.close()

            result = two_phase_analyze(
                combined_text,
                project_description=project_description,
                prompts=prompts,
            )
            if result.get("error"):
                session.status = "failed"
                session.error_message = "; ".join(result.get("warnings", ["分析失败"]))
                _update_db_session(session_id, "failed", session.error_message, 0, 0)
            else:
                session.functional_points = result["functional_points"]
                session.test_cases = result["test_cases"]
                session.status = "completed"
                if result.get("warnings"):
                    session.error_message = "; ".join(result["warnings"])
                _update_db_session(
                    session_id,
                    "completed",
                    session.error_message,
                    len(result["functional_points"]),
                    len(result["test_cases"]),
                    functional_points=result["functional_points"],
                    test_cases=result["test_cases"],
                )
        except Exception as e:
            logger.exception("Analysis failed")
            session.status = "failed"
            session.error_message = str(e)
            _update_db_session(session_id, "failed", str(e), 0, 0)

    thread = threading.Thread(target=_run_full_analysis, daemon=True)
    thread.start()

    return {"session_id": session_id, "status": "analyzing"}


def _update_db_session(session_id: str, status: str, error_msg: str, fp_count: int, tc_count: int,
                       functional_points: list = None, test_cases: list = None):
    """Update GenSession DB record after analysis completes, and persist results."""
    from app.database import SessionLocal
    db = SessionLocal()
    try:
        record = db.query(db_models.GenSession).filter(db_models.GenSession.id == session_id).first()
        if record:
            record.status = status
            record.error_message = error_msg
            record.functional_points_count = fp_count
            record.test_cases_count = tc_count
            if status in ("completed", "failed"):
                record.completed_at = datetime.now()

            if status == "completed" and functional_points and test_cases:
                for fp in functional_points:
                    db.add(db_models.GenFunctionalPoint(
                        session_id=session_id,
                        fp_id=fp.id,
                        module=fp.module,
                        name=fp.name,
                        description=fp.description,
                        category=fp.category,
                    ))
                for tc in test_cases:
                    db.add(db_models.GenTestCase(
                        session_id=session_id,
                        test_case_id=tc.test_case_id,
                        module=tc.module,
                        title=tc.title,
                        preconditions=tc.preconditions,
                        test_steps=tc.test_steps,
                        expected_result=tc.expected_result,
                        priority=tc.priority,
                    ))
            db.commit()
    finally:
        db.close()


@router.get("/status/{session_id}", response_model=GenStatusResponse)
async def get_status(session_id: str, user=Depends(get_current_user)):
    """Check analysis progress."""
    with _lock:
        session = _sessions.get(session_id)
    if not session:
        raise HTTPException(404, "Session not found")
    return GenStatusResponse(
        session_id=session.session_id,
        status=session.status,
        filename=session.filename,
        error_message=session.error_message,
        functional_points_count=len(session.functional_points),
        test_cases_count=len(session.test_cases),
    )


@router.get("/preview/{session_id}", response_model=GenPreviewResponse)
async def preview_results(session_id: str, user=Depends(get_current_user)):
    """Preview generated functional points and test cases."""
    with _lock:
        session = _sessions.get(session_id)
    if not session:
        raise HTTPException(404, "Session not found")
    if session.status != "completed":
        raise HTTPException(400, f"分析尚未完成，当前状态: {session.status}")

    fps = [
        {"id": fp.id, "module": fp.module, "name": fp.name, "category": fp.category, "description": fp.description}
        for fp in session.functional_points
    ]
    tcs = [
        GenPreviewItem(
            test_case_id=tc.test_case_id,
            module=tc.module,
            title=tc.title,
            preconditions=tc.preconditions,
            test_steps=tc.test_steps,
            expected_result=tc.expected_result,
            priority=tc.priority,
        )
        for tc in session.test_cases
    ]
    return GenPreviewResponse(
        session_id=session_id,
        functional_points=fps,
        test_cases=tcs,
    )


@router.post("/import", response_model=GenImportResponse)
async def import_test_cases(
    body: GenImportRequest,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Import selected test cases into a project."""
    from app.gen.models import TestCase as GenTestCaseModel

    test_cases_data = None

    with _lock:
        session = _sessions.get(body.session_id)
    if session and session.status == "completed":
        test_cases_data = session.test_cases
    else:
        record = db.query(db_models.GenSession).filter(db_models.GenSession.id == body.session_id).first()
        if not record:
            raise HTTPException(404, "记录不存在")
        if record.status != "completed":
            raise HTTPException(400, "分析尚未完成")
        db_tcs = db.query(db_models.GenTestCase).filter(
            db_models.GenTestCase.session_id == body.session_id
        ).all()
        test_cases_data = [
            GenTestCaseModel(
                test_case_id=tc.test_case_id,
                session_id=body.session_id,
                module=tc.module or "",
                title=tc.title or "",
                preconditions=tc.preconditions or "",
                test_steps=tc.test_steps or "",
                expected_result=tc.expected_result or "",
                priority=tc.priority or "中",
            )
            for tc in db_tcs
        ]

    project = db.query(db_models.Project).filter(db_models.Project.id == body.project_id).first()
    if not project:
        raise HTTPException(404, "项目不存在")

    from app.gen.adapter import import_test_cases as do_import
    created = do_import(db, body.project_id, test_cases_data, body.selected_ids)

    record = db.query(db_models.GenSession).filter(db_models.GenSession.id == body.session_id).first()
    if record:
        record.imported_count = (record.imported_count or 0) + len(created)
        if record.project_id is None:
            record.project_id = body.project_id
        db.commit()

    return GenImportResponse(
        imported_count=len(created),
        test_case_ids=[tc.id for tc in created],
    )


@router.get("/history", response_model=GenHistoryListResponse)
async def get_history(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    project_id: Optional[int] = Query(None, description="按项目筛选"),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Get analysis history list."""
    query = db.query(db_models.GenSession).order_by(db_models.GenSession.created_at.desc())
    if project_id is not None:
        query = query.filter(db_models.GenSession.project_id == project_id)
    total = query.count()
    items = query.offset((page - 1) * page_size).limit(page_size).all()

    return GenHistoryListResponse(
        items=[
            GenHistoryItem(
                id=item.id,
                filename=item.filename,
                filenames=json.loads(item.filenames) if item.filenames else [item.filename],
                project_id=item.project_id,
                project_name=item.project.name if item.project else "",
                project_description=item.project_description or "",
                status=item.status,
                error_message=item.error_message or "",
                functional_points_count=item.functional_points_count or 0,
                test_cases_count=item.test_cases_count or 0,
                imported_count=item.imported_count or 0,
                created_at=item.created_at,
                completed_at=item.completed_at,
            )
            for item in items
        ],
        total=total,
    )


@router.get("/history/{session_id}/export-xlsx")
async def export_gen_test_cases_xlsx(
    session_id: str,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Export generated test cases as xlsx file."""
    record = db.query(db_models.GenSession).filter(db_models.GenSession.id == session_id).first()
    if not record:
        raise HTTPException(404, "记录不存在")
    if record.status != "completed":
        raise HTTPException(400, f"分析未完成，状态: {record.status}")

    db_tcs = db.query(db_models.GenTestCase).filter(
        db_models.GenTestCase.session_id == session_id
    ).order_by(db_models.GenTestCase.id).all()

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # Header style
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["用例ID", "所属模块", "标题", "前置条件", "测试步骤", "预期结果", "优先级"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    cell_align = Alignment(vertical="top", wrap_text=True)
    for row_idx, tc in enumerate(db_tcs, 2):
        values = [
            tc.test_case_id,
            tc.module,
            tc.title,
            tc.preconditions or "",
            tc.test_steps or "",
            tc.expected_result or "",
            tc.priority,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = cell_align
            cell.border = thin_border

    # Column widths
    widths = [14, 16, 30, 24, 40, 40, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    output = wb.active  # use tempfile
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    filename = f"测试用例_{session_id[:8]}.xlsx"
    ascii_name = f"testcases_{session_id[:8]}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{ascii_name}"; filename*=UTF-8\'\'{quote(filename)}'},
    )
@router.get("/history/{session_id}", response_model=GenPreviewResponse)
async def get_history_detail(
    session_id: str,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Get analysis detail from DB."""
    record = db.query(db_models.GenSession).filter(db_models.GenSession.id == session_id).first()
    if not record:
        raise HTTPException(404, "记录不存在")
    if record.status != "completed":
        raise HTTPException(400, f"分析未完成，状态: {record.status}")

    db_fps = db.query(db_models.GenFunctionalPoint).filter(
        db_models.GenFunctionalPoint.session_id == session_id
    ).order_by(db_models.GenFunctionalPoint.fp_id).all()

    db_tcs = db.query(db_models.GenTestCase).filter(
        db_models.GenTestCase.session_id == session_id
    ).order_by(db_models.GenTestCase.id).all()

    fps = [
        {"id": fp.fp_id, "module": fp.module, "name": fp.name, "category": fp.category, "description": fp.description}
        for fp in db_fps
    ]
    tcs = [
        GenPreviewItem(
            test_case_id=tc.test_case_id,
            module=tc.module,
            title=tc.title,
            preconditions=tc.preconditions or "",
            test_steps=tc.test_steps or "",
            expected_result=tc.expected_result or "",
            priority=tc.priority or "中",
        )
        for tc in db_tcs
    ]
    return GenPreviewResponse(
        session_id=session_id,
        functional_points=fps,
        test_cases=tcs,
    )


@router.delete("/history/{session_id}")
async def delete_history(
    session_id: str,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Delete analysis history record."""
    record = db.query(db_models.GenSession).filter(db_models.GenSession.id == session_id).first()
    if not record:
        raise HTTPException(404, "记录不存在")

    # Also remove from in-memory if present
    with _lock:
        _sessions.pop(session_id, None)

    db.delete(record)
    db.commit()
    return {"message": "删除成功"}


class GenTestCaseUpdate(BaseModel):
    module: Optional[str] = None
    title: Optional[str] = None
    preconditions: Optional[str] = None
    test_steps: Optional[str] = None
    expected_result: Optional[str] = None
    priority: Optional[str] = None


@router.put("/history/{session_id}/test-cases/{test_case_id}")
async def update_gen_test_case(
    session_id: str,
    test_case_id: str,
    body: GenTestCaseUpdate,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Update a test case in the analysis session."""
    record = db.query(db_models.GenSession).filter(db_models.GenSession.id == session_id).first()
    if not record:
        raise HTTPException(404, "记录不存在")

    tc = db.query(db_models.GenTestCase).filter(
        db_models.GenTestCase.session_id == session_id,
        db_models.GenTestCase.test_case_id == test_case_id,
    ).first()
    if not tc:
        raise HTTPException(404, "用例不存在")

    if body.module is not None:
        tc.module = body.module
    if body.title is not None:
        tc.title = body.title
    if body.preconditions is not None:
        tc.preconditions = body.preconditions
    if body.test_steps is not None:
        tc.test_steps = body.test_steps
    if body.expected_result is not None:
        tc.expected_result = body.expected_result
    if body.priority is not None:
        tc.priority = body.priority

    db.commit()
    return {"message": "更新成功"}


@router.delete("/history/{session_id}/test-cases/{test_case_id}")
async def delete_gen_test_case(
    session_id: str,
    test_case_id: str,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Delete a test case from the analysis session."""
    record = db.query(db_models.GenSession).filter(db_models.GenSession.id == session_id).first()
    if not record:
        raise HTTPException(404, "记录不存在")

    tc = db.query(db_models.GenTestCase).filter(
        db_models.GenTestCase.session_id == session_id,
        db_models.GenTestCase.test_case_id == test_case_id,
    ).first()
    if not tc:
        raise HTTPException(404, "用例不存在")

    db.delete(tc)
    record.test_cases_count = max(0, (record.test_cases_count or 1) - 1)
    db.commit()
    return {"message": "删除成功"}


