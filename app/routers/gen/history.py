"""``/api/gen/history`` family of endpoints — they all read from (or mutate)
the persistent ``GenSession`` table.  In-memory cleanup of the matching
session is also done here for parity with the original behavior.
"""
from __future__ import annotations

import json
from io import BytesIO
from typing import Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from sqlalchemy.ext.asyncio import AsyncSession

from ... import crud
from ...auth import get_current_user, get_user_project_filter
from ...database import get_async_db
from .schemas import (
    GenHistoryItem,
    GenHistoryListResponse,
    GenPreviewItem,
    GenPreviewResponse,
    GenTestCaseUpdate,
)
from app.gen.cancel import CANCEL_MESSAGE
from .state import (
    _lock,
    _sessions,
    clear_gen_runtime,
    request_cancel_gen,
)
from .storage import delete_session_files, load_session_files, session_has_uploads

router = APIRouter()


def _session_can_retry(record) -> bool:
    if record.status not in ("failed", "cancelled"):
        return False
    return session_has_uploads(record.id)


def _check_session_ownership(record, user):
    """非管理员只能访问自己创建的会话，返回 None 表示通过；否则返回 HTTPException。"""
    if user.role == "admin":
        return None
    if record.user_id is None or record.user_id != user.id:
        raise HTTPException(403, "无权限访问该会话")
    return None


@router.get("/history", response_model=GenHistoryListResponse)
async def get_history(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    project_id: Optional[int] = Query(None, description="按项目筛选"),
    case_kind: Optional[str] = Query(None, description="按用例类型筛选：ui / functional"),
    db: AsyncSession = Depends(get_async_db),
    user=Depends(get_current_user),
) -> GenHistoryListResponse:
    """Get analysis history list."""

    # 项目权限检查：非管理员只能查看自己有权限的项目
    allowed_project_ids = get_user_project_filter(user)
    if allowed_project_ids is not None and project_id is not None:
        if project_id not in allowed_project_ids:
            raise HTTPException(403, "无权限访问该项目")

    # 非管理员只能看到自己的会话
    user_id_filter = None if user.role == "admin" else user.id
    kind = case_kind if case_kind in ("ui", "functional") else None
    result = await crud.gen.list_gen_sessions(
        db, page=page, page_size=page_size,
        project_id=project_id, user_id_filter=user_id_filter,
        case_kind=kind,
    )
    items = result["items"]
    total = result["total"]

    return GenHistoryListResponse(
        items=[
            GenHistoryItem(
                id=item.id,
                filename=item.filename,
                filenames=json.loads(item.filenames) if item.filenames else [item.filename],
                project_id=item.project_id,
                project_name=item.project.name if item.project else "",
                project_description=item.project_description or "",
                status=item.status,
                error_message=item.error_message or "",
                progress=int(item.progress or 0),
                progress_message=item.progress_message or "",
                functional_points_count=item.functional_points_count or 0,
                test_cases_count=item.test_cases_count or 0,
                imported_count=item.imported_count or 0,
                case_kind=(getattr(item, "case_kind", None) or "ui"),
                created_at=item.created_at,
                completed_at=item.completed_at,
                can_retry=_session_can_retry(item),
            )
            for item in items
        ],
        total=total,
    )


@router.get("/history/{session_id}/export-xlsx")
async def export_gen_test_cases_xlsx(
    session_id: str,
    db: AsyncSession = Depends(get_async_db),
    user=Depends(get_current_user),
) -> Response:
    """Export generated test cases as xlsx file."""

    record = await crud.gen.get_gen_session(db, session_id)
    if not record:
        raise HTTPException(404, "记录不存在")
    _check_session_ownership(record, user)
    if record.status != "completed":
        raise HTTPException(400, f"分析未完成，状态: {record.status}")

    db_tcs = await crud.gen.list_gen_test_cases(db, session_id)

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # Header style
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["用例ID", "所属模块", "标题", "前置条件", "测试步骤", "预期结果", "优先级"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    cell_align = Alignment(vertical="top", wrap_text=True)
    from app.gen.adapter import format_priority_for_export
    for row_idx, tc in enumerate(db_tcs, 2):
        values = [
            tc.test_case_id,
            tc.module,
            tc.title,
            tc.preconditions or "",
            tc.test_steps or "",
            tc.expected_result or "",
            format_priority_for_export(tc.priority),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = cell_align
            cell.border = thin_border

    # Column widths
    widths = [14, 16, 30, 24, 40, 40, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    output = wb.active  # use tempfile
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"测试用例_{session_id[:8]}.xlsx"
    ascii_name = f"testcases_{session_id[:8]}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{ascii_name}"; filename*=UTF-8\'\'{quote(filename)}'},
    )


@router.get("/history/{session_id}", response_model=GenPreviewResponse)
async def get_history_detail(
    session_id: str,
    db: AsyncSession = Depends(get_async_db),
    user=Depends(get_current_user),
) -> GenPreviewResponse:
    """Get analysis detail from DB."""

    record = await crud.gen.get_gen_session(db, session_id)
    if not record:
        raise HTTPException(404, "记录不存在")
    _check_session_ownership(record, user)
    # cancelled/failed may still have partial results worth reviewing

    db_fps = await crud.gen.list_gen_functional_points(db, session_id)
    db_tcs = await crud.gen.list_gen_test_cases(db, session_id)

    import json as _json
    filenames: list[str] = []
    if record.filenames:
        try:
            filenames = _json.loads(record.filenames)
        except Exception:
            filenames = [record.filename] if record.filename else []
    elif record.filename:
        filenames = [record.filename]

    fps = [
        {"id": fp.fp_id, "module": fp.module, "name": fp.name, "category": fp.category, "description": fp.description}
        for fp in db_fps
    ]
    tcs = [
        GenPreviewItem(
            test_case_id=tc.test_case_id,
            module=tc.module or "",
            title=tc.title,
            preconditions=tc.preconditions or "",
            test_steps=tc.test_steps or "",
            expected_result=tc.expected_result or "",
            priority=tc.priority or "中",
            selected=not bool(getattr(tc, "validation_errors", None)),
            validation_errors=getattr(tc, "validation_errors", None) or "",
            structured_steps=list(tc.structured_steps or [])
            if isinstance(getattr(tc, "structured_steps", None), list)
            else [],
        )
        for tc in db_tcs
    ]
    return GenPreviewResponse(
        session_id=session_id,
        filename=record.filename or "",
        filenames=filenames if isinstance(filenames, list) else [],
        status=record.status or "",
        error_message=record.error_message or "",
        progress=int(record.progress or 0),
        progress_message=record.progress_message or "",
        functional_points_count=record.functional_points_count or len(fps),
        test_cases_count=record.test_cases_count or len(tcs),
        case_kind=(getattr(record, "case_kind", None) or "ui"),
        project_id=record.project_id,
        project_name=record.project.name if getattr(record, "project", None) else "",
        functional_points=fps,
        test_cases=tcs,
    )


@router.delete("/history/{session_id}")
async def delete_history(
    session_id: str,
    db: AsyncSession = Depends(get_async_db),
    user=Depends(get_current_user),
) -> dict:
    """Delete analysis history record.

    In-progress sessions (``analyzing`` / ``pending``) must be stopped first.
    """

    record = await crud.gen.get_gen_session(db, session_id)
    if not record:
        raise HTTPException(404, "记录不存在")
    _check_session_ownership(record, user)

    if record.status in ("analyzing", "pending"):
        raise HTTPException(400, "分析进行中，请先停止后再删除")

    # Also remove from in-memory if present
    async with _lock:
        _sessions.pop(session_id, None)
    await clear_gen_runtime(session_id)

    await crud.gen.delete_gen_session(db, session_id)
    delete_session_files(session_id)
    return {"message": "删除成功"}


@router.post("/history/{session_id}/cancel")
async def cancel_analysis(
    session_id: str,
    db: AsyncSession = Depends(get_async_db),
    user=Depends(get_current_user),
) -> dict:
    """Stop an in-progress analysis (same worker cancels the asyncio task)."""
    from datetime import datetime

    record = await crud.gen.get_gen_session(db, session_id)
    if not record:
        raise HTTPException(404, "记录不存在")
    _check_session_ownership(record, user)

    if record.status not in ("analyzing", "pending"):
        raise HTTPException(400, f"当前状态无法停止：{record.status}")

    await request_cancel_gen(session_id)

    async with _lock:
        mem = _sessions.get(session_id)
        if mem:
            mem.status = "cancelled"
            mem.error_message = CANCEL_MESSAGE
            mem.progress_message = CANCEL_MESSAGE
            mem.progress = 100

    await crud.gen.update_gen_session_status(
        db,
        session_id,
        status="cancelled",
        error_message=CANCEL_MESSAGE,
        completed_at=datetime.now(),
    )
    return {"message": "已停止分析", "status": "cancelled"}


@router.post("/history/{session_id}/retry")
async def retry_analysis(
    session_id: str,
    agent_id: Optional[int] = Query(None, description="可选：指定生成 Agent"),
    db: AsyncSession = Depends(get_async_db),
    user=Depends(get_current_user),
) -> dict:
    """Re-run analysis for a failed/cancelled session using persisted uploads."""
    from io import BytesIO

    from app.crud.agent_definition import get_active_by_type, get_agent_definition
    from app.gen.models import AnalysisSession
    from app.gen.prompts import min_tcs_per_item, pick_fp_prompt_key, pick_tc_prompt_key, case_kind_from_tc_prompt_key
    from .upload import launch_gen_analysis

    record = await crud.gen.get_gen_session(db, session_id)
    if not record:
        raise HTTPException(404, "记录不存在")
    _check_session_ownership(record, user)

    if record.status in ("analyzing", "pending"):
        raise HTTPException(400, "分析进行中，请先停止后再重试")
    if record.status not in ("failed", "cancelled"):
        raise HTTPException(400, f"当前状态不可重试：{record.status}")

    loaded = load_session_files(session_id)
    if not loaded:
        raise HTTPException(400, "原始上传文件已丢失，请重新上传分析")
    filenames, raw_bytes = loaded

    selected_agent = None
    if agent_id is not None:
        selected_agent = await get_agent_definition(db, agent_id)
        if selected_agent is None or selected_agent.agent_type != "generation":
            raise HTTPException(400, f"无效的生成 Agent id={agent_id}")
    else:
        selected_agent = await get_active_by_type(db, "generation")

    resolved_agent_id = selected_agent.id if selected_agent else None
    resolved_skills = list(selected_agent.skills or []) if selected_agent else []
    fp_prompt_key = pick_fp_prompt_key(resolved_skills)
    tc_prompt_key = pick_tc_prompt_key(resolved_skills)
    min_tcs = min_tcs_per_item(resolved_skills, tc_prompt_key=tc_prompt_key)
    case_kind = case_kind_from_tc_prompt_key(tc_prompt_key)

    await clear_gen_runtime(session_id)
    async with _lock:
        _sessions.pop(session_id, None)

    await crud.gen.clear_gen_session_results(db, session_id)

    session = AnalysisSession(
        session_id=session_id,
        filename=filenames[0] if filenames else (record.filename or "unknown"),
        filenames=filenames,
        project_description=record.project_description or "",
        status="analyzing",
        progress=0,
        progress_message="准备重新分析",
        case_kind=case_kind,
    )
    async with _lock:
        _sessions[session_id] = session

    record.case_kind = case_kind
    record.status = "analyzing"
    await db.commit()

    file_contents = [BytesIO(b) for b in raw_bytes]
    await launch_gen_analysis(
        session_id=session_id,
        session=session,
        file_contents=file_contents,
        filenames=filenames,
        project_description=record.project_description or "",
        selected_agent=selected_agent,
        resolved_agent_id=resolved_agent_id,
        resolved_skills=resolved_skills,
        fp_prompt_key=fp_prompt_key,
        tc_prompt_key=tc_prompt_key,
        min_tcs=min_tcs,
    )
    return {
        "session_id": session_id,
        "status": "analyzing",
        "agent_id": resolved_agent_id,
        "tc_prompt_key": tc_prompt_key,
        "message": "已开始重新分析",
    }


@router.put("/history/{session_id}/test-cases/{test_case_id}")
async def update_gen_test_case(
    session_id: str,
    test_case_id: str,
    body: GenTestCaseUpdate,
    db: AsyncSession = Depends(get_async_db),
    user=Depends(get_current_user),
) -> dict:
    """Update a test case in the analysis session."""

    record = await crud.gen.get_gen_session(db, session_id)
    if not record:
        raise HTTPException(404, "记录不存在")
    _check_session_ownership(record, user)

    tc = await crud.gen.get_gen_test_case(db, session_id, test_case_id)
    if not tc:
        raise HTTPException(404, "用例不存在")

    await crud.gen.update_gen_test_case(db, session_id, test_case_id, body)
    return {"message": "更新成功"}


@router.delete("/history/{session_id}/test-cases/{test_case_id}")
async def delete_gen_test_case(
    session_id: str,
    test_case_id: str,
    db: AsyncSession = Depends(get_async_db),
    user=Depends(get_current_user),
) -> dict:
    """Delete a test case from the analysis session."""

    record = await crud.gen.get_gen_session(db, session_id)
    if not record:
        raise HTTPException(404, "记录不存在")
    _check_session_ownership(record, user)

    tc = await crud.gen.get_gen_test_case(db, session_id, test_case_id)
    if not tc:
        raise HTTPException(404, "用例不存在")

    await crud.gen.delete_gen_test_case(db, session_id, test_case_id)
    return {"message": "删除成功"}
