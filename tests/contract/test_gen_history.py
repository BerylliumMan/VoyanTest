# tests/contract/test_gen_history.py
"""``/api/gen/history*`` 契约测试。

覆盖目标:
- 列表分页 + project_id 过滤
- 单条详情 (404, 400 未完成, 200 已完成)
- 详情 xlsx 导出 (404, 400 未完成, 200 已完成)
- 删除 (404, 200, 内存 _sessions 同步清理)
- 测试用例更新 (404 record, 404 tc, 200 单字段/多字段, full update)
- 测试用例删除 (404, 200 计数 -1)
"""
from __future__ import annotations

import json
from datetime import datetime

import pytest
import pytest_asyncio
from sqlalchemy import select
from openpyxl import load_workbook

from app import db_models
from app.routers.gen import state
from app.database import Base


@pytest_asyncio.fixture(autouse=True)
async def _reset_state_and_db(db):
    """每个测试独立清理 ``_sessions`` 与相关表,避免测试间串扰。"""
    state._sessions.clear()
    state._cancelled_sessions.clear()
    state._gen_tasks.clear()
    yield
    state._sessions.clear()
    state._cancelled_sessions.clear()
    state._gen_tasks.clear()
    try:
        for table in (
            db_models.GenTestCase.__table__,
            db_models.GenFunctionalPoint.__table__,
            db_models.GenSession.__table__,
        ):
            await db.execute(table.delete())
        await db.commit()
    except Exception:
        await db.rollback()


# ==================== 辅助 ====================

async def _make_session(db, status="completed", project_id=None, filenames=None, **overrides):
    import uuid
    sid = overrides.pop("id", str(uuid.uuid4()))
    filenames = filenames or ["a.md"]
    record = db_models.GenSession(
        id=sid,
        filename=filenames[0],
        filenames=json.dumps(filenames),
        project_id=project_id,
        project_description=overrides.pop("project_description", "pdesc"),
        status=status,
        error_message=overrides.pop("error_message", None),
        functional_points_count=overrides.pop("functional_points_count", 0),
        test_cases_count=overrides.pop("test_cases_count", 0),
        imported_count=overrides.pop("imported_count", 0),
    )
    db.add(record)
    await db.commit()
    await db.refresh(record)
    return record


async def _make_fp(db, session_id, fp_id=1, module="m", name="n", description="d", category="c"):
    fp = db_models.GenFunctionalPoint(
        session_id=session_id,
        fp_id=fp_id,
        module=module,
        name=name,
        description=description,
        category=category,
    )
    db.add(fp)
    await db.commit()
    await db.refresh(fp)
    return fp


async def _make_tc(db, session_id, test_case_id="TC-1", module="m", title="t",
              preconditions="pre", test_steps="1. step 1", expected_result="pass",
              priority="中"):
    tc = db_models.GenTestCase(
        session_id=session_id,
        test_case_id=test_case_id,
        module=module,
        title=title,
        preconditions=preconditions,
        test_steps=test_steps,
        expected_result=expected_result,
        priority=priority,
    )
    db.add(tc)
    await db.commit()
    await db.refresh(tc)
    return tc


# ==================== 列表 ====================

class TestHistoryList:
    """GET /api/gen/history"""

    @pytest.mark.asyncio
    async def test_list_requires_auth(self, client):
        resp = client.get("/api/gen/history")
        assert resp.status_code == 401

    @pytest.mark.asyncio
    async def test_list_empty(self, client, admin_cookies):
        resp = client.get("/api/gen/history", cookies=admin_cookies)
        assert resp.status_code == 200
        data = resp.json()
        assert data["items"] == []
        assert data["total"] == 0

    @pytest.mark.asyncio
    async def test_list_with_records(self, client, admin_cookies, db):
        await _make_session(db, status="completed")
        await _make_session(db, status="failed")
        await _make_session(db, status="analyzing")
        resp = client.get("/api/gen/history", cookies=admin_cookies)
        assert resp.status_code == 200
        data = resp.json()
        assert data["total"] == 3
        assert len(data["items"]) == 3

    @pytest.mark.asyncio
    async def test_list_pagination(self, client, admin_cookies, db):
        for _ in range(5):
            await _make_session(db)
        resp = client.get("/api/gen/history?page=1&page_size=2", cookies=admin_cookies)
        assert resp.status_code == 200
        data = resp.json()
        assert data["total"] == 5
        assert len(data["items"]) == 2
        assert data["items"][0]["id"] != data["items"][1]["id"]

    @pytest.mark.asyncio
    async def test_list_filter_by_project_id(self, client, admin_cookies, db, sample_project):
        pid = sample_project["id"]
        await _make_session(db, project_id=pid)
        await _make_session(db, project_id=None)
        resp = client.get(f"/api/gen/history?project_id={pid}", cookies=admin_cookies)
        assert resp.status_code == 200
        data = resp.json()
        assert data["total"] == 1
        assert data["items"][0]["project_id"] == pid

    @pytest.mark.asyncio
    async def test_list_filenames_fallback_when_null(self, client, admin_cookies, db):
        """filenames 字段为 NULL 时回退到 [filename]。"""
        record = db_models.GenSession(
            id="sid-fb",
            filename="only.md",
            filenames=None,
            status="completed",
        )
        db.add(record)
        await db.commit()
        resp = client.get("/api/gen/history", cookies=admin_cookies)
        assert resp.status_code == 200
        item = next(x for x in resp.json()["items"] if x["id"] == "sid-fb")
        assert item["filenames"] == ["only.md"]


# ==================== 详情 ====================

class TestHistoryDetail:
    """GET /api/gen/history/{session_id}"""

    @pytest.mark.asyncio
    async def test_detail_requires_auth(self, client, db):
        rec = await _make_session(db)
        resp = client.get(f"/api/gen/history/{rec.id}")
        assert resp.status_code == 401

    @pytest.mark.asyncio
    async def test_detail_not_found(self, client, admin_cookies):
        resp = client.get("/api/gen/history/nonexistent", cookies=admin_cookies)
        assert resp.status_code == 404

    @pytest.mark.asyncio
    async def test_detail_not_completed_returns_400(self, client, admin_cookies, db):
        rec = await _make_session(db, status="analyzing")
        resp = client.get(f"/api/gen/history/{rec.id}", cookies=admin_cookies)
        assert resp.status_code == 400
        assert "analyzing" in resp.json()["detail"]

    @pytest.mark.asyncio
    async def test_detail_completed(self, client, admin_cookies, db):
        rec = await _make_session(db, status="completed", test_cases_count=2, functional_points_count=1)
        await _make_fp(db, rec.id, fp_id=1, module="登录", name="用户登录")
        await _make_tc(db, rec.id, test_case_id="TC-A", module="登录", title="登录")
        await _make_tc(db, rec.id, test_case_id="TC-B", module="登录", title="登出")
        resp = client.get(f"/api/gen/history/{rec.id}", cookies=admin_cookies)
        assert resp.status_code == 200
        data = resp.json()
        assert data["session_id"] == rec.id
        assert len(data["functional_points"]) == 1
        assert len(data["test_cases"]) == 2
        assert data["functional_points"][0]["name"] == "用户登录"
        assert {tc["test_case_id"] for tc in data["test_cases"]} == {"TC-A", "TC-B"}


# ==================== XLSX 导出 ====================

class TestHistoryExportXlsx:
    """GET /api/gen/history/{session_id}/export-xlsx"""

    @pytest.mark.asyncio
    async def test_export_requires_auth(self, client, db):
        rec = await _make_session(db, status="completed")
        resp = client.get(f"/api/gen/history/{rec.id}/export-xlsx")
        assert resp.status_code == 401

    @pytest.mark.asyncio
    async def test_export_not_found(self, client, admin_cookies):
        resp = client.get("/api/gen/history/nope/export-xlsx", cookies=admin_cookies)
        assert resp.status_code == 404

    @pytest.mark.asyncio
    async def test_export_not_completed_returns_400(self, client, admin_cookies, db):
        rec = await _make_session(db, status="analyzing")
        resp = client.get(f"/api/gen/history/{rec.id}/export-xlsx", cookies=admin_cookies)
        assert resp.status_code == 400
        assert "analyzing" in resp.json()["detail"]

    @pytest.mark.asyncio
    async def test_export_xlsx_returns_workbook(self, client, admin_cookies, db):
        rec = await _make_session(db, status="completed", test_cases_count=2)
        await _make_tc(db, rec.id, test_case_id="TC-1", title="用例一",
                 preconditions=None, test_steps="1. open 2. click", expected_result="ok", priority="高")
        await _make_tc(db, rec.id, test_case_id="TC-2", title="用例二",
                 preconditions="登录态", test_steps="1. logout", expected_result="bye", priority="中")
        resp = client.get(f"/api/gen/history/{rec.id}/export-xlsx", cookies=admin_cookies)
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith("application/vnd.openxmlformats")
        cd = resp.headers["content-disposition"]
        assert "attachment" in cd
        assert rec.id[:8] in cd

        wb = load_workbook(filename=__import__("io").BytesIO(resp.content))
        ws = wb.active
        assert ws.title == "测试用例"
        headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
        assert headers == ["用例ID", "所属模块", "标题", "前置条件", "测试步骤", "预期结果", "优先级"]
        assert ws.cell(row=2, column=1).value == "TC-1"
        assert ws.cell(row=2, column=3).value == "用例一"
        assert ws.cell(row=2, column=5).value == "1. open 2. click"
        assert ws.cell(row=2, column=4).value == "" or ws.cell(row=2, column=4).value is None
        assert ws.cell(row=3, column=1).value == "TC-2"
        assert ws.cell(row=3, column=4).value == "登录态"

    @pytest.mark.asyncio
    async def test_export_xlsx_empty_testcases(self, client, admin_cookies, db):
        rec = await _make_session(db, status="completed", test_cases_count=0)
        resp = client.get(f"/api/gen/history/{rec.id}/export-xlsx", cookies=admin_cookies)
        assert resp.status_code == 200
        wb = load_workbook(filename=__import__("io").BytesIO(resp.content))
        ws = wb.active
        assert ws.max_row == 1  # 仅 header


# ==================== 删除 ====================

class TestHistoryDelete:
    """DELETE /api/gen/history/{session_id}"""

    @pytest.mark.asyncio
    async def test_delete_requires_auth(self, client, db):
        rec = await _make_session(db)
        resp = client.delete(f"/api/gen/history/{rec.id}")
        assert resp.status_code == 401

    @pytest.mark.asyncio
    async def test_delete_not_found(self, client, admin_cookies):
        resp = client.delete("/api/gen/history/nope")
        assert resp.status_code == 404

    @pytest.mark.asyncio
    async def test_delete_success(self, client, admin_cookies, db):
        rec = await _make_session(db)
        resp = client.delete(f"/api/gen/history/{rec.id}", cookies=admin_cookies)
        assert resp.status_code == 200
        assert "成功" in resp.json()["message"]
        # DB 已删除
        record_result = await db.execute(
            select(db_models.GenSession).where(db_models.GenSession.id == rec.id)
        )
        assert record_result.scalar_one_or_none() is None

    @pytest.mark.asyncio
    async def test_delete_also_clears_in_memory_session(self, client, admin_cookies, db):
        """删除时同步清理 _sessions 中的内存对象。"""
        rec = await _make_session(db)
        from app.gen import models as gen_models
        state._sessions[rec.id] = gen_models.AnalysisSession(
            session_id=rec.id,
            filename=rec.filename,
            status="completed",
        )
        assert rec.id in state._sessions
        resp = client.delete(f"/api/gen/history/{rec.id}", cookies=admin_cookies)
        assert resp.status_code == 200
        assert rec.id not in state._sessions

    @pytest.mark.asyncio
    async def test_delete_silently_skips_missing_in_memory(self, client, admin_cookies, db):
        """内存中不存在时,_sessions.pop(..., None) 不报错。"""
        rec = await _make_session(db)
        state._sessions.pop(rec.id, None)
        resp = client.delete(f"/api/gen/history/{rec.id}", cookies=admin_cookies)
        assert resp.status_code == 200


# ==================== 停止分析 ====================

class TestHistoryCancel:
    """POST /api/gen/history/{session_id}/cancel"""

    @pytest.mark.asyncio
    async def test_cancel_requires_auth(self, client, db):
        rec = await _make_session(db, status="analyzing")
        resp = client.post(f"/api/gen/history/{rec.id}/cancel")
        assert resp.status_code == 401

    @pytest.mark.asyncio
    async def test_cancel_not_found(self, client, admin_cookies):
        resp = client.post("/api/gen/history/nope/cancel", cookies=admin_cookies)
        assert resp.status_code == 404

    @pytest.mark.asyncio
    async def test_cancel_analyzing_success(self, client, admin_cookies, db):
        rec = await _make_session(db, status="analyzing")
        resp = client.post(f"/api/gen/history/{rec.id}/cancel", cookies=admin_cookies)
        assert resp.status_code == 200
        assert resp.json()["status"] == "cancelled"
        await db.refresh(rec)
        assert rec.status == "cancelled"
        assert "停止" in (rec.error_message or "")

    @pytest.mark.asyncio
    async def test_cancel_completed_rejected(self, client, admin_cookies, db):
        rec = await _make_session(db, status="completed")
        resp = client.post(f"/api/gen/history/{rec.id}/cancel", cookies=admin_cookies)
        assert resp.status_code == 400


# ==================== 测试用例更新 ====================

class TestUpdateTestCase:
    """PUT /api/gen/history/{sid}/test-cases/{tc_id}"""

    @pytest.mark.asyncio
    async def test_update_requires_auth(self, client, db):
        rec = await _make_session(db, status="completed")
        await _make_tc(db, rec.id)
        resp = client.put(
            f"/api/gen/history/{rec.id}/test-cases/TC-1",
            json={"title": "x"},
        )
        assert resp.status_code == 401

    @pytest.mark.asyncio
    async def test_update_session_not_found(self, client, admin_cookies):
        resp = client.put(
            "/api/gen/history/nope/test-cases/TC-1",
            json={"title": "x"},
        )
        assert resp.status_code == 404

    @pytest.mark.asyncio
    async def test_update_testcase_not_found(self, client, admin_cookies, db):
        rec = await _make_session(db, status="completed")
        resp = client.put(
            f"/api/gen/history/{rec.id}/test-cases/TC-NOT-EXIST",
            json={"title": "x"},
        )
        assert resp.status_code == 404

    @pytest.mark.asyncio
    async def test_update_single_field(self, client, admin_cookies, db):
        rec = await _make_session(db, status="completed")
        tc = await _make_tc(db, rec.id, test_case_id="TC-U", title="old")
        resp = client.put(
            f"/api/gen/history/{rec.id}/test-cases/TC-U",
            json={"title": "new title"},
            cookies=admin_cookies,
        )
        assert resp.status_code == 200
        await db.refresh(tc)
        assert tc.title == "new title"

    @pytest.mark.asyncio
    async def test_update_multiple_fields(self, client, admin_cookies, db):
        rec = await _make_session(db, status="completed")
        tc = await _make_tc(db, rec.id, test_case_id="TC-M")
        body = {
            "module": "M2",
            "title": "T2",
            "preconditions": "P2",
            "test_steps": "1. S2",
            "expected_result": "R2",
            "priority": "高",
        }
        resp = client.put(
            f"/api/gen/history/{rec.id}/test-cases/TC-M",
            json=body,
            cookies=admin_cookies,
        )
        assert resp.status_code == 200
        await db.refresh(tc)
        assert tc.module == "M2"
        assert tc.title == "T2"
        assert tc.preconditions == "P2"
        assert tc.test_steps == "1. S2"
        assert tc.expected_result == "R2"
        assert tc.priority == "高"

    @pytest.mark.asyncio
    async def test_update_empty_body_keeps_values(self, client, admin_cookies, db):
        """body 中所有字段均为 None 时,任何字段都不会被改。"""
        rec = await _make_session(db, status="completed")
        tc = await _make_tc(db, rec.id, test_case_id="TC-K", title="original")
        resp = client.put(
            f"/api/gen/history/{rec.id}/test-cases/TC-K",
            json={},
            cookies=admin_cookies,
        )
        assert resp.status_code == 200
        await db.refresh(tc)
        assert tc.title == "original"


# ==================== 测试用例删除 ====================

class TestDeleteTestCase:
    """DELETE /api/gen/history/{sid}/test-cases/{tc_id}"""

    @pytest.mark.asyncio
    async def test_delete_tc_requires_auth(self, client, db):
        rec = await _make_session(db, status="completed")
        await _make_tc(db, rec.id)
        resp = client.delete(f"/api/gen/history/{rec.id}/test-cases/TC-1")
        assert resp.status_code == 401

    @pytest.mark.asyncio
    async def test_delete_tc_session_not_found(self, client, admin_cookies):
        resp = client.delete("/api/gen/history/nope/test-cases/TC-1")
        assert resp.status_code == 404

    @pytest.mark.asyncio
    async def test_delete_tc_not_found(self, client, admin_cookies, db):
        rec = await _make_session(db, status="completed")
        resp = client.delete(f"/api/gen/history/{rec.id}/test-cases/TC-X")
        assert resp.status_code == 404

    @pytest.mark.asyncio
    async def test_delete_tc_decrements_counter(self, client, admin_cookies, db):
        rec = await _make_session(db, status="completed", test_cases_count=3)
        await _make_tc(db, rec.id, test_case_id="TC-D")
        resp = client.delete(
            f"/api/gen/history/{rec.id}/test-cases/TC-D",
            cookies=admin_cookies,
        )
        assert resp.status_code == 200
        assert "成功" in resp.json()["message"]
        await db.refresh(rec)
        assert rec.test_cases_count == 2
        tc_result = await db.execute(
            select(db_models.GenTestCase).where(
                db_models.GenTestCase.test_case_id == "TC-D"
            )
        )
        assert tc_result.scalar_one_or_none() is None

    @pytest.mark.asyncio
    async def test_delete_tc_counter_floor_zero(self, client, admin_cookies, db):
        """counter 已经 0 时,减 1 仍保持 0 (max(0, -1))。"""
        rec = await _make_session(db, status="completed", test_cases_count=0)
        await _make_tc(db, rec.id, test_case_id="TC-Z")
        rec.test_cases_count = 0
        await db.commit()
        resp = client.delete(
            f"/api/gen/history/{rec.id}/test-cases/TC-Z",
            cookies=admin_cookies,
        )
        assert resp.status_code == 200
        await db.refresh(rec)
        assert rec.test_cases_count == 0
